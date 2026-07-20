import os

import sys

import math

import logging

from openpyxl import load_workbook, Workbook

from openpyxl.utils import get_column_letter

from openpyxl.styles import Font, PatternFill

import argparse

from datetime import datetime

import random

 

LOGS_DIRNAME = 'logs'

SPECIMEN_FINAL      = 'Specimen Final'

SPECIMEN            = 'Specimen'

ORIGINATOR          = 'Originator'

SSTI                = 'SSTI'

BLOOD               = 'Blood'

RESULTS             = 'Results'

MRSA                = 'MRSA'

MSSA                = 'MSSA'

PVL                 = 'RT-pvl Result'

POSITIVE            = 'Positive'

NEGATIVE            = 'Negative'

 

MRSA_POSITIVE_SSTI      = 'MRSA PVL POSITIVE SSTI'

MRSA_POSITIVE_BLOOD     = 'MRSA PVL POSITIVE BLOOD'

MRSA_NEGATIVE_SSTI      = 'MRSA PVL NEGATIVE SSTI'

MRSA_NEGATIVE_BLOOD     = 'MRSA PVL NEGATIVE BLOOD'

 

ORANGE_FILL = PatternFill(start_color="FFB732", end_color="FFB732", fill_type="solid")

GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

 

TITLES = [MRSA_POSITIVE_SSTI, MRSA_POSITIVE_BLOOD, MRSA_NEGATIVE_SSTI, MRSA_NEGATIVE_BLOOD, MSSA]

 

def setup_logs():

    logs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), LOGS_DIRNAME)

 

    timestamp = datetime.now().strftime("%y%m%d_%H%M%S")

 

    logs_file = os.path.join(logs_dir, f'logs__{timestamp}.txt')

       

    os.makedirs(logs_dir, exist_ok=True)

   

    logging.basicConfig(

        filename=logs_file,

        level=logging.DEBUG,

        format="%(message)s"

    )

 

def display_and_log(msg, is_error=False):

    print(f'{msg}\n')

    if is_error:

        logging.error(f'{msg}\n')

    else:

        logging.info(f'{msg}\n')

 

def load_data(input_file, sheetname):

    """Load data from Excel file into a list of dictionaries"""

    if not (input_file.endswith(".xls") or input_file.endswith(".xlsx")):

        raise ValueError("Unsupported file format. Please provide Excel file.")

   

    wb = load_workbook(input_file, read_only=True, data_only=True)

   

    if sheetname not in wb.sheetnames:

        raise ValueError(f"Sheet '{sheetname}' not found in the workbook")

   

    ws = wb[sheetname]

   

    headers = []

    data = []

   

    for i, row in enumerate(ws.rows):

        if i == 0:  # First row is headers

            headers = [cell.value for cell in row]

        else:

            values = [cell.value for cell in row]

            row_data = dict(zip(headers, values))

            data.append(row_data)

   

    return data

 

def determine_result(row):

    '''

    if RT-mecA Result or RT-mecA Result positive = MRSA, if both negative = MSSA

    '''

    if row["RT-mecA Result"] == POSITIVE or row["RT-mecC Result"] == POSITIVE:

        return MRSA

    return MSSA

 

from openpyxl import Workbook

 

def export_data_to_excel(data, headers, output_file, sheetname="Sheet1"):

    """Export data to a specific sheet in Excel, create the file if it doesn't exist."""

   

    # Load workbook if it exists, otherwise create a new one

    if os.path.exists(output_file):

        wb = load_workbook(output_file)

    else:

        wb = Workbook()

        # Remove default sheet only if a new sheet will be added

        default_sheet = wb.active

        if default_sheet.title == "Sheet":

            wb.remove(default_sheet)

 

    # Create new sheet or overwrite if exists

    if sheetname in wb.sheetnames:

        del wb[sheetname]  # remove if you want to replace it

   

    ws = wb.create_sheet(title=sheetname)

 

    # Write header

    for col_idx, header in enumerate(headers, 1):

        ws.cell(row=1, column=col_idx, value=header)

 

    # Write data

    for row_idx, row in enumerate(data, 2):

        for col_idx, value in enumerate(row, 1):

            ws.cell(row=row_idx, column=col_idx, value=value)

 

    wb.save(output_file)

    return wb

 

def order_the_table(data, columns):

    """Order the data into sections based on criteria"""

    final_data = []

   

    def add_section(title, subset_data):

        """Helper function to add a title, column headers, the data, and an empty row."""

        if subset_data:

            title_row = [title] + [''] * (len(columns) - 1)

            header_row = columns

            empty_row = [''] * len(columns)

           

            final_data.append(title_row)

            final_data.append(header_row)

            final_data.extend(subset_data)

            final_data.append(empty_row)

   

    # Convert data to rows for each section

    mssa_rows = []

    mrsa_pos_ssti_rows = []

    mrsa_pos_blood_rows = []

    mrsa_neg_ssti_rows = []

    mrsa_neg_blood_rows = []

   

    for row_dict in data:

        row_values = [row_dict.get(col, '') for col in columns]

        if row_dict[RESULTS] == MSSA:

            mssa_rows.append(row_values)

        elif row_dict[RESULTS] == MRSA and row_dict[PVL] == POSITIVE and row_dict[SPECIMEN_FINAL] == SSTI:

            mrsa_pos_ssti_rows.append(row_values)

        elif row_dict[RESULTS] == MRSA and row_dict[PVL] == POSITIVE and row_dict[SPECIMEN_FINAL] == BLOOD:

            mrsa_pos_blood_rows.append(row_values)

        elif row_dict[RESULTS] == MRSA and row_dict[PVL] == NEGATIVE and row_dict[SPECIMEN_FINAL] == SSTI:

            mrsa_neg_ssti_rows.append(row_values)

        elif row_dict[RESULTS] == MRSA and row_dict[PVL] == NEGATIVE and row_dict[SPECIMEN_FINAL] == BLOOD:

            mrsa_neg_blood_rows.append(row_values)


    # Add sections in order

    add_section(MSSA, mssa_rows)

    add_section(MRSA_POSITIVE_SSTI, mrsa_pos_ssti_rows)

    add_section(MRSA_POSITIVE_BLOOD, mrsa_pos_blood_rows)

    add_section(MRSA_NEGATIVE_SSTI, mrsa_neg_ssti_rows)

    add_section(MRSA_NEGATIVE_BLOOD, mrsa_neg_blood_rows)

   

    return final_data

 

def style_excel(output_path, titles, data, sheetname="Sheet1"):

    """Apply formatting: bold titles, dodger blue headers, and add 3 colored values after the title."""

    wb = load_workbook(output_path)

 

    if sheetname in wb.sheetnames:

        ws = wb[sheetname]

    else:

        ws = wb.active  # fallback

 

    # Dynamic column width adjustment

    column_widths = {}

    for row in ws.iter_rows():

        for cell in row:

            if cell.value:

                col_letter = get_column_letter(cell.column)

                column_widths[col_letter] = max(column_widths.get(col_letter, 0), len(str(cell.value)))

 

    for col_letter, width in column_widths.items():

        ws.column_dimensions[col_letter].width = min(width + 5, 50)

 

    bold_font = Font(bold=True)

    color_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")  # Light Steel Blue

 

    # Add values after each title

    for row_idx, row in enumerate(ws.iter_rows(), start=1):

        first_cell = row[0]

        title = first_cell.value

 

        if title in titles:

            # Apply bold style to title row

            for cell in row:

                cell.font = bold_font

           

            # Color the next row (header)

            if row_idx + 1 <= ws.max_row:

                header_row = ws[row_idx + 1]

                for cell in header_row:

                    cell.fill = color_fill

 

            # Add 3 values with pastel fills

            title_column = first_cell.column

           

            if title in data:

                ws.cell(row=row_idx, column=title_column + 1, value=data[title]['needed']).fill = GREEN_FILL

                ws.cell(row=row_idx, column=title_column + 2, value=data[title]['found']).fill = YELLOW_FILL

                ws.cell(row=row_idx, column=title_column + 3, value=data[title]['added']).fill = ORANGE_FILL

 

    wb.save(output_path)

 

def highlight_added_rows(file_path, rows_added, sheetname='Sheet1'):

    """

    Processes an Excel file to find tables under specific titles and highlight extra rows.

    """

    wb = load_workbook(file_path)

    ws = wb.active

 

    if sheetname in wb.sheetnames:

        ws = wb[sheetname]

    else:

        ws = wb.active

   

    # Convert worksheet to list of lists for easier processing

    data = []

    for row in ws.values:

        data.append(list(row))

   

    for i, row in enumerate(data):

        if row and row[0] in TITLES and row[0] in rows_added:

            title = row[0]

            row_limit = rows_added[title]

            table_start = i + 1  # Table starts below title

           

            # Find the table end (empty row marks the end)

            table_end = table_start

            while table_end < len(data) and any(val is not None and val != '' for val in data[table_end]):

                table_end += 1

               

            table_size = table_end - table_start  # Number of rows in table

           

            # Highlight rows that exceed the row_limit

            if table_size > row_limit:

                for idx in range(table_start + row_limit, table_end):

                    for col in range(1, ws.max_column + 1):

                        ws.cell(row=idx + 1, column=col).fill = ORANGE_FILL  # OpenPyXL uses 1-based indexing

   

    wb.save(file_path)

 

def select_by_originator(data, column_list, originator_key, num_needed):

    """Select rows with proportionate distribution by originator"""

    if num_needed <= 0:

        return []

   

    # Count occurrences of each originator

    originator_counts = {}

    for row in data:

        orig = row[originator_key]

        if orig not in originator_counts:

            originator_counts[orig] = 0

        originator_counts[orig] += 1

   

    total_count = sum(originator_counts.values())

    selected_rows = []

   

    # Calculate how many to select from each originator

    for originator, count in originator_counts.items():

        fraction = count / total_count

        num_from_originator = max(1, round(fraction * num_needed))

       

        # Get all rows for this originator

        originator_rows = [row for row in data if row[originator_key] == originator]

        # Randomly sample rows
        if num_from_originator < len(originator_rows):
            # Set seed for reproducibility
            random.seed(42)
            sampled_rows = random.sample(originator_rows, min(num_from_originator, len(originator_rows)))
        else:
            sampled_rows = originator_rows

        selected_rows.extend(sampled_rows)

    # Ensure we don't return more than needed
    return selected_rows[:num_needed]

 
def create_freezing_test_result(df_data, df):
    # Extract sample numbers
    samples_data = {entry['Sample No'] for entry in df_data}
    samples_df = {entry['Sample No'] for entry in df}

    # Build result list
    results = []

    for sample in samples_data:
        results.append({
            'Sample Name': sample,
            'Test Name': 'Freezing',
            'Test Result': 'yes' if sample in samples_df else 'no'
        })

    return results

def dict_list_to_rows(dict_list, columns):
    """Convert list of dicts to list of rows (list of lists) based on given column order."""
    return [[entry.get(col, "") for col in columns] for entry in dict_list]

 

def select_specimens(data, output_file, num_of_specimen=30):
    # Create columns list from the first row's keys
    columns = list(data[0].keys())
    bact_freeze_sheetname = 'ToFreeze'

    # Add SPECIMEN_FINAL column to each row
    specimen_values_to_replace = ['Wound', 'Abscess', 'Surgery wound', 'Skin', 'Skin ulcer', 'Ear',
                                 'Ear L', 'Ear R', 'Elbow', 'Elbow L', 'Elbow R', 'Nose',
                                 'Lesion', 'Eye', 'Nasal wash', 'Navel', 'Perineum']

    for row in data:
        # Create SPECIMEN_FINAL column with the same value as SPECIMEN
        row[SPECIMEN_FINAL] = row[SPECIMEN]

        # Replace specified values with SSTI
        if row[SPECIMEN_FINAL] in specimen_values_to_replace:
            row[SPECIMEN_FINAL] = SSTI

        # Fix results based on RT-mecA and RT-mecC
        row[RESULTS] = determine_result(row)

        # Ensure ORIGINATOR is a string
        row[ORIGINATOR] = str(row[ORIGINATOR])

    # Create results directory if it doesn't exist
    os.makedirs("results", exist_ok=True)

    timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
    output_filename = f"{output_file}/bact_to_freeze__{timestamp}.xlsx"

    # First, select all MRSA PVL Positive (priority)
    matched_data = [row for row in data if
                    row[RESULTS] == MRSA and
                    row[PVL] == POSITIVE and
                    row[SPECIMEN_FINAL] in [SSTI, BLOOD]]

    matched_count = len(matched_data)
    remaining_needed = num_of_specimen - matched_count

    display_and_log(f"[+] Found {matched_count} matching MRSA PVL positive specimens. Need {remaining_needed} more.")

    # Remaining data excluding matched
    remaining_data = [row for row in data if not(
                      row[RESULTS] == MRSA and
                      row[PVL] == POSITIVE and
                      row[SPECIMEN_FINAL] in [SSTI, BLOOD])]

    # Filter data for different categories
    ssti_data = [row for row in remaining_data if
                row[RESULTS] == MRSA and
                row[PVL] == NEGATIVE and
                row[SPECIMEN_FINAL] == SSTI]

    blood_data = [row for row in remaining_data if
                row[RESULTS] == MRSA and
                row[PVL] == NEGATIVE and
                row[SPECIMEN_FINAL] == BLOOD]


    mssa_positive_data = [row for row in remaining_data if
                        row[RESULTS] == MSSA and
                        row[PVL] == POSITIVE and
                        row[SPECIMEN_FINAL] in [SSTI, BLOOD]]

    mssa_negative_data = [row for row in remaining_data if
                        row[RESULTS] == MSSA and
                        row[PVL] == NEGATIVE and
                        row[SPECIMEN_FINAL] in [SSTI, BLOOD]]

    # Calculate how many of each type needed
    mrsa_negative_ssti_needed = math.ceil((remaining_needed * 0.5))
    mrsa_negative_blood_needed = len(blood_data)  # All MRSA PVL Negative Blood specimens
    mssa_needed = math.floor((remaining_needed * 0.25))

    display_and_log(f"[+] Total Needed : ")
    display_and_log(f"    - MRSA PVL NEGATIVE SSTI  :  {mrsa_negative_ssti_needed}.")
    display_and_log(f"    - MRSA PVL NEGATIVE BLOOD :  {mrsa_negative_blood_needed}.")
    display_and_log(f"    - MSSA PVL NEG / POS      : {mssa_needed}.")

    # Select specimens by originator
    selected_ssti = select_by_originator(ssti_data, columns, ORIGINATOR, mrsa_negative_ssti_needed)
    selected_blood = blood_data  # Take all MRSA PVL Negative Blood specimens

    ssti_fetched_len_on_start = len(selected_ssti)
    blood_fetched_len_on_start = len(selected_blood)

    mrsa_negative_ssti_needed_before_completion = 0
    mrsa_negative_blood_needed_before_completion = 0
    mssa_needed_before_completion = 0

    is_found_less_ssti_than_needed = mrsa_negative_ssti_needed > ssti_fetched_len_on_start
    is_found_less_blood_than_needed = mrsa_negative_blood_needed > blood_fetched_len_on_start

    # Handle case where not enough SSTI specimens
    if is_found_less_ssti_than_needed and not is_found_less_blood_than_needed:
        display_and_log(f"[!] MRSA-NEG-SSTI : Found {ssti_fetched_len_on_start} on {mrsa_negative_ssti_needed} needed : Completion with more MRSA-NEG-BLOOD ")

        mrsa_negative_blood_needed_before_completion = mrsa_negative_blood_needed
        mrsa_negative_blood_complement = mrsa_negative_ssti_needed - ssti_fetched_len_on_start
        mrsa_negative_blood_needed += mrsa_negative_blood_complement

        selected_blood = select_by_originator(blood_data, columns, ORIGINATOR, mrsa_negative_blood_needed)

   
    # Handle case where not enough BLOOD specimens
    elif not is_found_less_ssti_than_needed and is_found_less_blood_than_needed:
        display_and_log(f"[!] MRSA-NEG-BLOOD : Found {blood_fetched_len_on_start} on {mrsa_negative_blood_needed} needed : Completion with more MRSA-NEG-SSTI ")

        mrsa_negative_ssti_needed_before_completion = mrsa_negative_ssti_needed
        mrsa_negative_ssti_complement = mrsa_negative_blood_needed - blood_fetched_len_on_start
        mrsa_negative_ssti_needed += mrsa_negative_ssti_complement

        selected_ssti = select_by_originator(ssti_data, columns, ORIGINATOR, mrsa_negative_ssti_needed)

    # Combine selected data
    df_selection = matched_data + selected_ssti + selected_blood

   

    # If we still need more specimens, get them from MSSA
    selected_mssa_positive = []
    selected_mssa_negative = []

    if len(df_selection) < num_of_specimen:
        mssa_needed_before_completion = mssa_needed
        mssa_needed = num_of_specimen - len(df_selection)

   
        # First try to get MSSA PVL positive
        selected_mssa_positive = select_by_originator(
            mssa_positive_data, columns, ORIGINATOR,
            min(mssa_needed, len(mssa_positive_data))
        )

   
        # If more MSSA specimens needed, select PVL negative ones
        remaining_mssa_needed = mssa_needed - len(selected_mssa_positive)

        selected_mssa_negative = select_by_originator(
            mssa_negative_data, columns, ORIGINATOR,
            remaining_mssa_needed
        )

        display_and_log(f"[+] MSSA needed (needed + completion) total {mssa_needed}: ")
        display_and_log(f"    - PVL + : Found {len(selected_mssa_positive)}.")
        display_and_log(f"    - PVL - : Found {len(selected_mssa_negative)}.")

        df_selection = matched_data + selected_ssti + selected_blood + selected_mssa_positive + selected_mssa_negative

    # Order the data for output
    df_ordered = order_the_table(df_selection, columns)

    # Export to Excel
    export_data_to_excel(df_ordered, columns, output_filename, bact_freeze_sheetname)

    # Create test result sheet
    result_data = create_freezing_test_result(data, df_selection)
    columns = ['Sample Name', 'Test Name', 'Test Result']
    rows = dict_list_to_rows(result_data, columns)
    export_data_to_excel(rows, columns, output_filename, sheetname='TestResults')

 

    # Add data next to each title table
    mssa_len = len(selected_mssa_positive) + len(selected_mssa_negative)
    data_mssa_needed = mssa_needed_before_completion
    data_mssa_found = mssa_needed_before_completion if mssa_needed_before_completion < mssa_len else mssa_len
    data_mssa_added = mssa_len - mssa_needed_before_completion if mssa_needed_before_completion > 0 else 0

    ssti_len = len(selected_ssti)
    data_mrsa_negative_ssti_needed = mrsa_negative_ssti_needed if mrsa_negative_ssti_needed_before_completion == 0 else mrsa_negative_ssti_needed_before_completion
    data_mrsa_negative_ssti_found = ssti_len if mrsa_negative_ssti_needed_before_completion == 0 else ssti_fetched_len_on_start
    data_mrsa_negative_ssti_added = ssti_len - mrsa_negative_ssti_needed_before_completion if mrsa_negative_ssti_needed_before_completion > 0 else 0

    blood_len = len(selected_blood)
    data_mrsa_negative_blood_needed = mrsa_negative_blood_needed if mrsa_negative_blood_needed_before_completion == 0 else mrsa_negative_blood_needed_before_completion
    data_mrsa_negative_blood_found = blood_len if mrsa_negative_blood_needed_before_completion == 0 else blood_fetched_len_on_start
    data_mrsa_negative_blood_added = blood_len - mrsa_negative_blood_needed_before_completion if mrsa_negative_blood_needed_before_completion > 0 else 0

    data = {
        MSSA: {
            'needed': f'Needed {data_mssa_needed}',
            'found': f'Found {data_mssa_found}',
            'added': f'Added {data_mssa_added} to reach {num_of_specimen}',

        },

        MRSA_NEGATIVE_SSTI: {
            'needed': f'Needed {data_mrsa_negative_ssti_needed}',
            'found': f'Found {data_mrsa_negative_ssti_found}',
            'added': f'Added {data_mrsa_negative_ssti_added} to reach {num_of_specimen}',

        },

        MRSA_NEGATIVE_BLOOD: {
            'needed': f'Needed {data_mrsa_negative_blood_needed}',
            'found': f'Found {data_mrsa_negative_blood_found}',
            'added': f'Added {data_mrsa_negative_blood_added} to reach {num_of_specimen}',
        }
    }

    style_excel(output_filename, TITLES, data, bact_freeze_sheetname)

   

    # Highlight added rows
    rows_added = {}

    if mssa_needed_before_completion > 0:
        rows_added[MSSA] = mssa_needed_before_completion + 1

    if mrsa_negative_blood_needed_before_completion > 0:
        rows_added[MRSA_NEGATIVE_BLOOD] = mrsa_negative_blood_needed_before_completion + 1

   
    if mrsa_negative_ssti_needed_before_completion > 0:
        rows_added[MRSA_NEGATIVE_SSTI] = mrsa_negative_ssti_needed_before_completion + 1

    if rows_added:  
        display_and_log(f"[+] Highlighting complementions rows.")
        highlight_added_rows(output_filename, rows_added, bact_freeze_sheetname)

    return output_filename


def main():
    parser = argparse.ArgumentParser(description="Select specimens to freeze from an input file.")
    parser.add_argument("input_file", type=str, help="Path to input Excel file")
    parser.add_argument("output_dir", type=str, help="Path to output dir")
    parser.add_argument("sheetname", type=str, help="sheetname")
    parser.add_argument("num_of_specimen", type=int, help="Number of specimens to select")

    args = parser.parse_args()
    num_of_specimen = args.num_of_specimen

    setup_logs()

    display_and_log(f"\n\n[>>>] Generate the freeze bact file ({num_of_specimen} specimens)\n")

    df_data = load_data(args.input_file, args.sheetname)

    output_dir = args.output_dir

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    select_specimens(df_data, output_dir, num_of_specimen)

    display_and_log(f"[+] File saved successfully to {output_dir}.")
    display_and_log(f"\n[---] The End ")

 
if __name__ == "__main__":
    main()
